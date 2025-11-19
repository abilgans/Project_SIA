# Main.py
"""
Project SIA - Kelompok Ngapak
Single-file desktop app (Tkinter) with OTP activation on Register.
Includes automated cycle and full financial reports (Income Statement, Statement of Changes in Equity, Balance Sheet).
Python 3.14+
"""
import os
import sqlite3
import hashlib
import smtplib
import random
import threading
import time
from email.mime.text import MIMEText
from datetime import datetime, timedelta
from decimal import Decimal, getcontext
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog

# Optional libs (graceful fallback)
try:
    import pandas as pd
except Exception:
    pd = None

try:
    import matplotlib
    matplotlib.use("TkAgg")
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    import matplotlib.pyplot as plt
except Exception:
    matplotlib = None
    FigureCanvasTkAgg = None
    plt = None

try:
    from openpyxl import Workbook
except Exception:
    Workbook = None

try:
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
except Exception:
    SimpleDocTemplate = None

getcontext().prec = 28

# ---------------- CONFIG ----------------
DB_FILENAME = "akuntansi.db"
APP_TITLE = "Project SIA - Kelompok Ngapak"

WINDOW_BG = "#F8FAFC"
COLOR_PRIMARY = "#0077B6"
COLOR_ACCENT = "#00B4D8"
COLOR_TEXT = "#023E8A"
CARD_BG = "#FFFFFF"
FONT = ("Segoe UI", 10)

# SMTP (untuk OTP) - set sendiri jika mau
SMTP_EMAIL = "siangapak@gmail.com"
SMTP_PASSWORD = "pwwc mdza lbtu mvtz"
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
OTP_TTL_SECONDS = 600

# ---------------- Utilities ----------------
def hash_password(pw: str) -> str:
    return hashlib.sha256(pw.encode('utf-8')).hexdigest()

def to_decimal(x):
    try:
        return Decimal(str(x))
    except Exception:
        return Decimal('0')

def moneyfmt(x):
    try:
        d = to_decimal(x)
        return f"{d:,.2f}"
    except Exception:
        return str(x)

# ---------------- Database ----------------
def init_db_if_needed():
    need_init = not os.path.exists(DB_FILENAME)
    conn = sqlite3.connect(DB_FILENAME)
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            is_active INTEGER DEFAULT 0,
            otp TEXT,
            otp_expiry TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS akun (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            no_akun TEXT UNIQUE NOT NULL,
            nama_akun TEXT NOT NULL,
            tipe TEXT NOT NULL
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS jurnal (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tanggal TEXT NOT NULL,
            akun_debit TEXT NOT NULL,
            akun_kredit TEXT NOT NULL,
            debit REAL DEFAULT 0,
            kredit REAL DEFAULT 0,
            keterangan TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS adjusting (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tanggal TEXT NOT NULL,
            akun_debit TEXT NOT NULL,
            akun_kredit TEXT NOT NULL,
            debit REAL DEFAULT 0,
            kredit REAL DEFAULT 0,
            keterangan TEXT
        )
    """)
    conn.commit()
    conn.close()
    return need_init

def get_conn():
    if not os.path.exists(DB_FILENAME):
        init_db_if_needed()
    return sqlite3.connect(DB_FILENAME)

# ---------------- User / Auth (activation OTP) ----------------
def user_exists(email):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id FROM users WHERE email=?", (email,))
    r = cur.fetchone()
    conn.close()
    return r is not None

def register_user_with_otp(email, password):
    if user_exists(email):
        return False, "Email sudah terdaftar."
    ph = hash_password(password)
    otp = f"{random.randint(100000,999999)}"
    expiry = (datetime.now() + timedelta(seconds=OTP_TTL_SECONDS)).isoformat()
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("INSERT INTO users (email, password_hash, is_active, otp, otp_expiry) VALUES (?, ?, 0, ?, ?)",
                (email, ph, otp, expiry))
    conn.commit()
    conn.close()
    sent, msg = send_activation_otp_email(email, otp)
    if sent:
        return True, "Terdaftar. OTP aktivasi terkirim ke email."
    else:
        print(f"[DEBUG] OTP for {email}: {otp} (expiry {expiry})")
        return True, f"Terdaftar. Gagal kirim email ({msg}), OTP ditampilkan di console untuk testing."

def send_activation_otp_email(to_email, otp):
    if not SMTP_EMAIL or not SMTP_PASSWORD:
        return False, "SMTP belum dikonfigurasi."
    try:
        body = f"Kode OTP aktivasi akun Anda: {otp}\nBerlaku selama {OTP_TTL_SECONDS//60} menit."
        msg = MIMEText(body)
        msg['Subject'] = "OTP Aktivasi Akun - Project SIA"
        msg['From'] = SMTP_EMAIL
        msg['To'] = to_email
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT, timeout=10)
        server.starttls()
        server.login(SMTP_EMAIL, SMTP_PASSWORD)
        server.sendmail(SMTP_EMAIL, [to_email], msg.as_string())
        server.quit()
        return True, "Terkirim"
    except Exception as e:
        return False, str(e)

def verify_activation_otp(email, otp_input):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT otp, otp_expiry FROM users WHERE email=?", (email,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return False, "Email tidak ditemukan."
    otp, expiry = row
    if not otp:
        conn.close()
        return False, "Tidak ada OTP tersimpan."
    try:
        exp_dt = datetime.fromisoformat(expiry)
    except Exception:
        exp_dt = datetime.now() - timedelta(seconds=1)
    if datetime.now() > exp_dt:
        conn.close()
        return False, "OTP sudah kadaluarsa."
    if str(otp_input).strip() != str(otp).strip():
        conn.close()
        return False, "OTP salah."
    cur.execute("UPDATE users SET is_active=1, otp=NULL, otp_expiry=NULL WHERE email=?", (email,))
    conn.commit()
    conn.close()
    return True, "Akun berhasil diaktifkan."

def verify_password_and_active(email, password):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT password_hash, is_active FROM users WHERE email=?", (email,))
    row = cur.fetchone()
    conn.close()
    if not row:
        return False, "Email tidak terdaftar."
    phash, is_active = row
    if phash != hash_password(password):
        return False, "Password salah."
    if is_active != 1:
        return False, "Akun belum aktif. Silakan cek email untuk OTP aktivasi."
    return True, "Login berhasil."

# ---------------- Akun & Jurnal CRUD ----------------
def list_accounts():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT no_akun, nama_akun, tipe FROM akun ORDER BY no_akun")
    rows = cur.fetchall()
    conn.close()
    return rows

def add_account_db(no, nama, tipe):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("INSERT INTO akun (no_akun, nama_akun, tipe) VALUES (?, ?, ?)", (no, nama, tipe))
    conn.commit()
    conn.close()

def edit_account_db(no, nama, tipe):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE akun SET nama_akun=?, tipe=? WHERE no_akun=?", (nama, tipe, no))
    conn.commit()
    conn.close()

def delete_account_db(no):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM jurnal WHERE akun_debit=? OR akun_kredit=?", (no, no))
    if cur.fetchone()[0] > 0:
        conn.close()
        raise ValueError("Tidak dapat menghapus akun yang memiliki transaksi.")
    cur.execute("DELETE FROM akun WHERE no_akun=?", (no,))
    conn.commit()
    conn.close()

def add_journal_db(tanggal, akun_debit, akun_kredit, debit, kredit, keterangan):
    if float(debit) < 0 or float(kredit) < 0:
        raise ValueError("Debit/Kredit tidak boleh negatif.")
    if float(debit) == 0 and float(kredit) == 0:
        raise ValueError("Isi debit atau kredit minimal satu.")
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM akun WHERE no_akun=?", (akun_debit,))
    if cur.fetchone()[0] == 0:
        conn.close()
        raise ValueError(f"Akun debit {akun_debit} tidak ditemukan")
    cur.execute("SELECT COUNT(*) FROM akun WHERE no_akun=?", (akun_kredit,))
    if cur.fetchone()[0] == 0:
        conn.close()
        raise ValueError(f"Akun kredit {akun_kredit} tidak ditemukan")
    cur.execute("""INSERT INTO jurnal (tanggal, akun_debit, akun_kredit, debit, kredit, keterangan)
                   VALUES (?, ?, ?, ?, ?, ?)""", (tanggal, akun_debit, akun_kredit, float(debit), float(kredit), keterangan))
    conn.commit()
    conn.close()

def list_journal_entries():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id, tanggal, akun_debit, akun_kredit, debit, kredit, keterangan FROM jurnal ORDER BY tanggal, id")
    rows = cur.fetchall()
    conn.close()
    return rows

def delete_journal_db(_id):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM jurnal WHERE id=?", (_id,))
    conn.commit()
    conn.close()

# Adjusting (penyesuaian) CRUD
def add_adjusting_db(tanggal, akun_debit, akun_kredit, debit, kredit, keterangan):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""INSERT INTO adjusting (tanggal, akun_debit, akun_kredit, debit, kredit, keterangan)
                   VALUES (?, ?, ?, ?, ?, ?)""", (tanggal, akun_debit, akun_kredit, float(debit), float(kredit), keterangan))
    conn.commit()
    conn.close()

def list_adjusting_entries():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id, tanggal, akun_debit, akun_kredit, debit, kredit, keterangan FROM adjusting ORDER BY tanggal, id")
    rows = cur.fetchall()
    conn.close()
    return rows

def delete_adjusting_db(_id):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM adjusting WHERE id=?", (_id,))
    conn.commit()
    conn.close()

def apply_adjustments():
    """
    Apply all adjusting entries by copying them into jurnal table with today's date
    and a note appended " (Penyesuaian from <tanggal>)". Does not delete adjusting rows.
    """
    adj = list_adjusting_entries()
    if not adj:
        return 0
    today = datetime.now().strftime("%Y-%m-%d")
    applied = 0
    conn = get_conn()
    cur = conn.cursor()
    for _id, tanggal, ad, ak, d, k, ket in adj:
        note = f"{ket} (Penyesuaian from {tanggal})"
        cur.execute("""INSERT INTO jurnal (tanggal, akun_debit, akun_kredit, debit, kredit, keterangan)
                       VALUES (?, ?, ?, ?, ?, ?)""", (today, ad, ak, float(d), float(k), note))
        applied += 1
    conn.commit()
    conn.close()
    return applied

# ---------------- Accounting computations ----------------
def compute_balances(include_adjustments=False):
    """
    include_adjustments: if True, compute balances using jurnal + adjusting entries (without applying)
    """
    accs = {}
    for no, nama, tipe in list_accounts():
        accs[no] = {'nama': nama, 'tipe': tipe, 'debit': to_decimal('0'), 'kredit': to_decimal('0')}
    # base jurnal
    for row in list_journal_entries():
        _id, tanggal, ad, ak, d, k, ket = row
        d = to_decimal(d)
        k = to_decimal(k)
        if ad not in accs:
            accs[ad] = {'nama': ad, 'tipe': 'Unknown', 'debit': to_decimal('0'), 'kredit': to_decimal('0')}
        if ak not in accs:
            accs[ak] = {'nama': ak, 'tipe': 'Unknown', 'debit': to_decimal('0'), 'kredit': to_decimal('0')}
        accs[ad]['debit'] += d
        accs[ak]['kredit'] += k
    if include_adjustments:
        for row in list_adjusting_entries():
            _id, tanggal, ad, ak, d, k, ket = row
            d = to_decimal(d)
            k = to_decimal(k)
            if ad not in accs:
                accs[ad] = {'nama': ad, 'tipe': 'Unknown', 'debit': to_decimal('0'), 'kredit': to_decimal('0')}
            if ak not in accs:
                accs[ak] = {'nama': ak, 'tipe': 'Unknown', 'debit': to_decimal('0'), 'kredit': to_decimal('0')}
            accs[ad]['debit'] += d
            accs[ak]['kredit'] += k
    return accs

def compute_trial_rows(include_adjustments=False):
    bal = compute_balances(include_adjustments=include_adjustments)
    rows = []
    for no in sorted(bal.keys()):
        d = bal[no]['debit']
        c = bal[no]['kredit']
        if d >= c:
            debit_bal = d - c
            credit_bal = to_decimal('0')
        else:
            debit_bal = to_decimal('0')
            credit_bal = c - d
        rows.append((no, bal[no]['nama'], bal[no]['tipe'], debit_bal, credit_bal))
    return rows

def prepare_balance_and_ratios(include_adjustments=False):
    rows = compute_trial_rows(include_adjustments=include_adjustments)
    assets = to_decimal('0')
    liabilities = to_decimal('0')
    equity = to_decimal('0')
    for no, nm, tipe, d, c in rows:
        bal = d - c
        t = (tipe or "").lower()
        if t in ('asset', 'aset'):
            assets += max(bal, to_decimal('0'))
        elif t in ('liability', 'kewajiban', 'liab'):
            liabilities += max(-bal, to_decimal('0'))
        elif t in ('equity', 'ekuitas'):
            equity += max(-bal, to_decimal('0'))
    current_ratio = None
    debt_to_equity = None
    try:
        if liabilities != to_decimal('0'):
            current_ratio = assets / liabilities
    except Exception:
        current_ratio = None
    try:
        if equity != to_decimal('0'):
            debt_to_equity = liabilities / equity
    except Exception:
        debt_to_equity = None
    return {'assets': assets, 'liabilities': liabilities, 'equity': equity,
            'current_ratio': current_ratio, 'debt_to_equity': debt_to_equity}

# ---------------- Financial statements ----------------
def compute_financial_statements(include_adjustments=False):
    balances = compute_balances(include_adjustments=include_adjustments)
    revenues = []
    expenses = []
    for no, info in balances.items():
        t = (info['tipe'] or "").lower()
        d = info['debit']
        c = info['kredit']
        bal = d - c
        if t in ('revenue', 'pendapatan', 'income'):
            rev_amt = max(to_decimal('0'), -bal)
            if rev_amt != 0:
                revenues.append((no, info['nama'], rev_amt))
        elif t in ('expense', 'beban', 'cost'):
            exp_amt = max(to_decimal('0'), bal)
            if exp_amt != 0:
                expenses.append((no, info['nama'], exp_amt))
    total_revenue = sum((r[2] for r in revenues), to_decimal('0'))
    total_expense = sum((e[2] for e in expenses), to_decimal('0'))
    net_income = total_revenue - total_expense

    assets = []
    liabilities = []
    equity_lst = []
    for no, info in balances.items():
        t = (info['tipe'] or "").lower()
        d = info['debit']
        c = info['kredit']
        bal = d - c
        if t in ('asset', 'aset'):
            amt = max(bal, to_decimal('0'))
            if amt != 0:
                assets.append((no, info['nama'], amt))
        elif t in ('liability', 'kewajiban', 'liab'):
            amt = max(-bal, to_decimal('0'))
            if amt != 0:
                liabilities.append((no, info['nama'], amt))
        elif t in ('equity', 'ekuitas'):
            amt = max(-bal, to_decimal('0'))
            if amt != 0:
                equity_lst.append((no, info['nama'], amt))

    total_equity_end = sum((e[2] for e in equity_lst), to_decimal('0'))
    total_equity_begin = total_equity_end - net_income
    equity_reconciliation = [
        ("Beginning Equity (approx)", total_equity_begin),
        ("Net Income (Laba/Rugi)", net_income),
        ("Ending Equity", total_equity_end)
    ]

    income_statement = {
        'revenues': revenues,
        'expenses': expenses,
        'total_revenue': total_revenue,
        'total_expense': total_expense,
        'net_income': net_income
    }
    balance_sheet = {
        'assets': assets,
        'liabilities': liabilities,
        'equity': equity_lst,
        'total_assets': sum((a[2] for a in assets), to_decimal('0')),
        'total_liabilities': sum((l[2] for l in liabilities), to_decimal('0')),
        'total_equity': total_equity_end
    }
    return income_statement, balance_sheet, equity_reconciliation

def export_reports_to_pdf(path, include_adjustments=True):
    if SimpleDocTemplate is None:
        raise RuntimeError("reportlab tidak terpasang")
    income, balance, eq_rec = compute_financial_statements(include_adjustments=include_adjustments)
    doc = SimpleDocTemplate(path, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Laporan Keuangan", styles['Title']))
    story.append(Spacer(1, 12))

    # Income Statement
    story.append(Paragraph("Laporan Laba Rugi", styles['Heading2']))
    data = [["Jenis", "No Akun", "Nama Akun", "Jumlah"]]
    for no, nm, amt in income['revenues']:
        data.append(["Pendapatan", no, nm, f"{float(amt):,.2f}"])
    data.append(["", "", "Total Pendapatan", f"{float(income['total_revenue']):,.2f}"])
    for no, nm, amt in income['expenses']:
        data.append(["Beban", no, nm, f"{float(amt):,.2f}"])
    data.append(["", "", "Total Beban", f"{float(income['total_expense']):,.2f}"])
    data.append(["", "", "Laba (Rugi) Bersih", f"{float(income['net_income']):,.2f}"])
    t = Table(data, colWidths=[80, 80, 220, 100])
    t.setStyle(TableStyle([('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                           ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                           ('ALIGN', (-1, 0), (-1, -1), 'RIGHT')]))
    story.append(t)
    story.append(Spacer(1, 12))

    # Equity reconciliation
    story.append(Paragraph("Perubahan Ekuitas (Rekonsiliasi sederhana)", styles['Heading2']))
    data = [["Keterangan", "Jumlah"]]
    for label, amt in eq_rec:
        data.append([label, f"{float(amt):,.2f}"])
    t2 = Table(data, colWidths=[400, 120])
    t2.setStyle(TableStyle([('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                            ('ALIGN', (-1, 0), (-1, -1), 'RIGHT')]))
    story.append(t2)
    story.append(Spacer(1, 12))

    # Balance sheet
    story.append(Paragraph("Neraca (Balance Sheet)", styles['Heading2']))
    data = [["Sisi", "No Akun", "Nama Akun", "Jumlah"]]
    for no, nm, amt in balance['assets']:
        data.append(["Aset", no, nm, f"{float(amt):,.2f}"])
    data.append(["", "", "Total Aset", f"{float(balance['total_assets']):,.2f}"])
    for no, nm, amt in balance['liabilities']:
        data.append(["Kewajiban", no, nm, f"{float(amt):,.2f}"])
    for no, nm, amt in balance['equity']:
        data.append(["Ekuitas", no, nm, f"{float(amt):,.2f}"])
    data.append(["", "", "Total Kewajiban + Ekuitas", f"{float(balance['total_liabilities'] + balance['total_equity']):,.2f}"])
    t3 = Table(data, colWidths=[80, 80, 220, 100])
    t3.setStyle(TableStyle([('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                            ('ALIGN', (-1, 0), (-1, -1), 'RIGHT')]))
    story.append(t3)

    doc.build(story)

# ---------------- Export helpers ----------------
def export_trial_to_excel(path, include_adjustments=False):
    if Workbook is None:
        raise RuntimeError("openpyxl tidak terpasang")
    rows = compute_trial_rows(include_adjustments)
    wb = Workbook()
    ws = wb.active
    ws.title = "Neraca Saldo"
    ws.append(["No Akun", "Nama", "Tipe", "Debit", "Kredit"])
    for r in rows:
        ws.append([r[0], r[1], r[2], float(r[3]), float(r[4])])
    wb.save(path)

def export_journal_to_excel(path):
    if Workbook is None:
        raise RuntimeError("openpyxl tidak terpasang")
    rows = list_journal_entries()
    wb = Workbook()
    ws = wb.active
    ws.title = "Jurnal"
    ws.append(["ID", "Tanggal", "Akun Debit", "Akun Kredit", "Debit", "Kredit", "Keterangan"])
    for r in rows:
        ws.append([r[0], r[1], r[2], r[3], float(r[4]), float(r[5]), r[6]])
    wb.save(path)

def export_adjusting_to_excel(path):
    if Workbook is None:
        raise RuntimeError("openpyxl tidak terpasang")
    rows = list_adjusting_entries()
    wb = Workbook()
    ws = wb.active
    ws.title = "Penyesuaian"
    ws.append(["ID", "Tanggal", "Akun Debit", "Akun Kredit", "Debit", "Kredit", "Keterangan"])
    for r in rows:
        ws.append([r[0], r[1], r[2], r[3], float(r[4]), float(r[5]), r[6]])
    wb.save(path)

def export_reports_to_excel(path, include_adjustments=True):
    if Workbook is None:
        raise RuntimeError("openpyxl tidak terpasang")
    income, balance, eq_rec = compute_financial_statements(include_adjustments=include_adjustments)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Laba Rugi"
    ws1.append(["Jenis", "No Akun", "Nama Akun", "Jumlah"])
    for no, nm, amt in income['revenues']:
        ws1.append(["Pendapatan", no, nm, float(amt)])
    ws1.append(["", "", "Total Pendapatan", float(income['total_revenue'])])
    for no, nm, amt in income['expenses']:
        ws1.append(["Beban", no, nm, float(amt)])
    ws1.append(["", "", "Total Beban", float(income['total_expense'])])
    ws1.append(["", "", "Laba (Rugi) Bersih", float(income['net_income'])])

    ws2 = wb.create_sheet("Perubahan Ekuitas")
    ws2.append(["Keterangan", "Jumlah"])
    for label, amt in eq_rec:
        ws2.append([label, float(amt)])

    ws3 = wb.create_sheet("Neraca")
    ws3.append(["Sisi", "No Akun", "Nama Akun", "Jumlah"])
    for no, nm, amt in balance['assets']:
        ws3.append(["Aset", no, nm, float(amt)])
    ws3.append(["", "", "Total Aset", float(balance['total_assets'])])
    for no, nm, amt in balance['liabilities']:
        ws3.append(["Kewajiban", no, nm, float(amt)])
    for no, nm, amt in balance['equity']:
        ws3.append(["Ekuitas", no, nm, float(amt)])
    ws3.append(["", "", "Total Kewajiban + Ekuitas", float(balance['total_liabilities'] + balance['total_equity'])])

    wb.save(path)

# ---------------- OTP cleanup (noop) ----------------
def _cleanup_worker():
    while True:
        time.sleep(60)

_cleanup_thread = threading.Thread(target=_cleanup_worker, daemon=True)
_cleanup_thread.start()

# ---------------- UI ----------------
class ActivationDialog(tk.Toplevel):
    def __init__(self, parent, email):
        super().__init__(parent)
        self.title("Aktivasi Akun")
        self.geometry("420x180")
        self.resizable(False, False)
        self.email = email
        ttk.Label(self, text=f"Masukkan kode OTP yang dikirim ke {email}").pack(padx=12, pady=(12, 6))
        self.entry_otp = ttk.Entry(self)
        self.entry_otp.pack(fill='x', padx=12)
        ttk.Button(self, text="Aktivasi", command=self.do_activate).pack(pady=12)
        ttk.Button(self, text="Tutup", command=self.destroy).pack()

    def do_activate(self):
        otp = self.entry_otp.get().strip()
        if not otp:
            messagebox.showwarning("Peringatan", "Masukkan OTP terlebih dahulu")
            return
        ok, msg = verify_activation_otp(self.email, otp)
        if ok:
            messagebox.showinfo("Sukses", msg)
            self.destroy()
        else:
            messagebox.showerror("Error", msg)

class LoginWindow(tk.Toplevel):
    def __init__(self, parent, on_success):
        super().__init__(parent)
        self.on_success = on_success
        self.overrideredirect(True)
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{sw}x{sh}+0+0")
        self.configure(bg=WINDOW_BG)
        self.build()

    def build(self):
        frm = ttk.Frame(self, padding=20)
        frm.place(relx=0.5, rely=0.5, anchor='center')
        title = tk.Label(frm, text=APP_TITLE, font=("Segoe UI", 20, "bold"), fg=COLOR_PRIMARY, bg=WINDOW_BG)
        title.pack(pady=(0, 12))
        nb = ttk.Notebook(frm)
        tab_login = ttk.Frame(nb)
        tab_reg = ttk.Frame(nb)
        nb.add(tab_login, text="Login")
        nb.add(tab_reg, text="Register")
        nb.pack(fill='both', expand=True)
        # Login
        ttk.Label(tab_login, text="Email:").pack(anchor='w', padx=8, pady=(8, 2))
        self.login_email = ttk.Entry(tab_login, width=40)
        self.login_email.pack(padx=8)
        ttk.Label(tab_login, text="Password:").pack(anchor='w', padx=8, pady=(8, 2))
        self.login_pass = ttk.Entry(tab_login, show='*', width=40)
        self.login_pass.pack(padx=8)
        ttk.Button(tab_login, text="Login", command=self.do_login).pack(pady=10)
        # Register
        ttk.Label(tab_reg, text="Email:").pack(anchor='w', padx=8, pady=(8, 2))
        self.reg_email = ttk.Entry(tab_reg, width=40)
        self.reg_email.pack(padx=8)
        ttk.Label(tab_reg, text="Password:").pack(anchor='w', padx=8, pady=(8, 2))
        self.reg_pass = ttk.Entry(tab_reg, show='*', width=40)
        self.reg_pass.pack(padx=8)
        ttk.Button(tab_reg, text="Register & Kirim OTP Aktivasi", command=self.do_register).pack(pady=10)
        ttk.Button(frm, text="Tutup (Esc)", command=self.quit_app).pack(pady=(8, 0))
        self.bind("<Escape>", lambda e: self.quit_app())

    def do_register(self):
        email = self.reg_email.get().strip()
        pw = self.reg_pass.get().strip()
        if not email or not pw:
            messagebox.showwarning("Peringatan", "Email & password wajib diisi")
            return
        ok, msg = register_user_with_otp(email, pw)
        if ok:
            messagebox.showinfo("Sukses", msg)
            ActivationDialog(self, email)
        else:
            messagebox.showerror("Error", msg)

    def do_login(self):
        email = self.login_email.get().strip()
        pw = self.login_pass.get().strip()
        if not email or not pw:
            messagebox.showwarning("Peringatan", "Email & password wajib diisi")
            return
        ok, msg = verify_password_and_active(email, pw)
        if ok:
            self.destroy()
            self.on_success(email)
        else:
            messagebox.showerror("Gagal login", msg)

    def quit_app(self):
        if messagebox.askyesno("Keluar", "Keluar aplikasi?"):
            self.master.destroy()

class AccountDialog(tk.Toplevel):
    def __init__(self, parent, initial=None, edit_mode=False):
        super().__init__(parent)
        self.title("Akun")
        self.geometry("420x240")
        self.configure(bg=WINDOW_BG)
        self.result = None
        self.edit_mode = edit_mode
        ttk.Label(self, text="No Akun:").pack(anchor='w', padx=12, pady=6)
        self.e_no = ttk.Entry(self)
        self.e_no.pack(fill='x', padx=12)
        ttk.Label(self, text="Nama Akun:").pack(anchor='w', padx=12, pady=6)
        self.e_nama = ttk.Entry(self)
        self.e_nama.pack(fill='x', padx=12)
        ttk.Label(self, text="Tipe (Asset/Liability/Equity/Revenue/Expense):").pack(anchor='w', padx=12, pady=6)
        self.combo = ttk.Combobox(self, values=["Asset", "Liability", "Equity", "Revenue", "Expense"])
        self.combo.pack(fill='x', padx=12)
        ttk.Button(self, text="Simpan", command=self.save).pack(pady=12)
        if initial:
            no, nama, tipe = initial
            self.e_no.insert(0, no)
            self.e_nama.insert(0, nama)
            self.combo.set(tipe)
            if edit_mode:
                self.e_no.config(state='disabled')

    def save(self):
        no = self.e_no.get().strip()
        nama = self.e_nama.get().strip()
        tipe = self.combo.get().strip()
        if not (no and nama and tipe):
            messagebox.showerror("Error", "Semua field wajib diisi")
            return
        self.result = (no, nama, tipe)
        self.destroy()

class JournalDialog(tk.Toplevel):
    def __init__(self, parent, accounts):
        super().__init__(parent)
        self.title("Tambah Jurnal")
        self.geometry("560x480")
        self.configure(bg=WINDOW_BG)
        self.result = None
        ttk.Label(self, text="Tanggal (YYYY-MM-DD):").pack(anchor='w', padx=12, pady=6)
        self.e_date = ttk.Entry(self)
        self.e_date.pack(fill='x', padx=12)
        self.e_date.insert(0, datetime.now().strftime("%Y-%m-%d"))
        ttk.Label(self, text="Akun Debit (no - nama):").pack(anchor='w', padx=12, pady=6)
        self.combo_debit = ttk.Combobox(self, values=[f"{a[0]} - {a[1]}" for a in accounts])
        self.combo_debit.pack(fill='x', padx=12)
        ttk.Label(self, text="Akun Kredit (no - nama):").pack(anchor='w', padx=12, pady=6)
        self.combo_credit = ttk.Combobox(self, values=[f"{a[0]} - {a[1]}" for a in accounts])
        self.combo_credit.pack(fill='x', padx=12)
        ttk.Label(self, text="Debit (angka, 0 jika tidak ada):").pack(anchor='w', padx=12, pady=6)
        self.e_debit = ttk.Entry(self)
        self.e_debit.pack(fill='x', padx=12)
        self.e_debit.insert(0, "0")
        ttk.Label(self, text="Kredit (angka, 0 jika tidak ada):").pack(anchor='w', padx=12, pady=6)
        self.e_kredit = ttk.Entry(self)
        self.e_kredit.pack(fill='x', padx=12)
        self.e_kredit.insert(0, "0")
        ttk.Label(self, text="Keterangan:").pack(anchor='w', padx=12, pady=6)
        self.e_ket = ttk.Entry(self)
        self.e_ket.pack(fill='x', padx=12)
        ttk.Button(self, text="Simpan", command=self.save).pack(pady=12)

    def save(self):
        tanggal = self.e_date.get().strip()
        dsel = self.combo_debit.get().strip()
        csel = self.combo_credit.get().strip()
        debit = self.e_debit.get().strip()
        kredit = self.e_kredit.get().strip()
        ket = self.e_ket.get().strip()
        if not (tanggal and dsel and csel):
            messagebox.showerror("Error", "Tanggal dan akun wajib diisi")
            return
        try:
            debit_val = float(debit.replace(",", "")) if debit else 0.0
            kredit_val = float(kredit.replace(",", "")) if kredit else 0.0
            adebit = dsel.split(" - ", 1)[0].strip()
            acredit = csel.split(" - ", 1)[0].strip()
            if debit_val < 0 or kredit_val < 0:
                raise ValueError("Nilai tidak boleh negatif")
            if debit_val == 0 and kredit_val == 0:
                raise ValueError("Isi debit atau kredit minimal satu")
            datetime.strptime(tanggal, "%Y-%m-%d")
            self.result = (tanggal, adebit, acredit, debit_val, kredit_val, ket)
            self.destroy()
        except Exception as e:
            messagebox.showerror("Error", str(e))

class AdjustDialog(tk.Toplevel):
    def __init__(self, parent, accounts):
        super().__init__(parent)
        self.title("Tambah Jurnal Penyesuaian")
        self.geometry("560x480")
        self.configure(bg=WINDOW_BG)
        self.result = None
        ttk.Label(self, text="Tanggal (YYYY-MM-DD):").pack(anchor='w', padx=12, pady=6)
        self.e_date = ttk.Entry(self)
        self.e_date.pack(fill='x', padx=12)
        self.e_date.insert(0, datetime.now().strftime("%Y-%m-%d"))
        ttk.Label(self, text="Akun Debit (no - nama):").pack(anchor='w', padx=12, pady=6)
        self.combo_debit = ttk.Combobox(self, values=[f"{a[0]} - {a[1]}" for a in accounts])
        self.combo_debit.pack(fill='x', padx=12)
        ttk.Label(self, text="Akun Kredit (no - nama):").pack(anchor='w', padx=12, pady=6)
        self.combo_credit = ttk.Combobox(self, values=[f"{a[0]} - {a[1]}" for a in accounts])
        self.combo_credit.pack(fill='x', padx=12)
        ttk.Label(self, text="Debit (angka, 0 jika tidak ada):").pack(anchor='w', padx=12, pady=6)
        self.e_debit = ttk.Entry(self)
        self.e_debit.pack(fill='x', padx=12)
        self.e_debit.insert(0, "0")
        ttk.Label(self, text="Kredit (angka, 0 jika tidak ada):").pack(anchor='w', padx=12, pady=6)
        self.e_kredit = ttk.Entry(self)
        self.e_kredit.pack(fill='x', padx=12)
        self.e_kredit.insert(0, "0")
        ttk.Label(self, text="Keterangan:").pack(anchor='w', padx=12, pady=6)
        self.e_ket = ttk.Entry(self)
        self.e_ket.pack(fill='x', padx=12)
        ttk.Button(self, text="Simpan Penyesuaian", command=self.save).pack(pady=12)

    def save(self):
        tanggal = self.e_date.get().strip()
        dsel = self.combo_debit.get().strip()
        csel = self.combo_credit.get().strip()
        debit = self.e_debit.get().strip()
        kredit = self.e_kredit.get().strip()
        ket = self.e_ket.get().strip()
        if not (tanggal and dsel and csel):
            messagebox.showerror("Error", "Tanggal dan akun wajib diisi")
            return
        try:
            debit_val = float(debit.replace(",", "")) if debit else 0.0
            kredit_val = float(kredit.replace(",", "")) if kredit else 0.0
            adebit = dsel.split(" - ", 1)[0].strip()
            acredit = csel.split(" - ", 1)[0].strip()
            if debit_val < 0 or kredit_val < 0:
                raise ValueError("Nilai tidak boleh negatif")
            if debit_val == 0 and kredit_val == 0:
                raise ValueError("Isi debit atau kredit minimal satu")
            datetime.strptime(tanggal, "%Y-%m-%d")
            self.result = (tanggal, adebit, acredit, debit_val, kredit_val, ket)
            self.destroy()
        except Exception as e:
            messagebox.showerror("Error", str(e))

# ---------------- Main Application ----------------
class MainApp(tk.Tk):
    def __init__(self, user_email):
        super().__init__()
        self.user_email = user_email
        self.overrideredirect(True)
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{sw}x{sh}+0+0")
        self._is_fullscreen = True
        self.configure(bg=WINDOW_BG)
        self.style = ttk.Style(self)
        try:
            self.style.theme_use('clam')
        except Exception:
            pass
        self.style.configure('TFrame', background=WINDOW_BG)
        self.style.configure('TLabel', background=WINDOW_BG, foreground=COLOR_TEXT, font=FONT)
        self.style.configure('Header.TLabel', font=("Segoe UI", 14, 'bold'), foreground=COLOR_PRIMARY, background=WINDOW_BG)
        self.build_header()
        self.build_notebook()
        self.after(150, self.refresh_all)
        self.bind("<F11>", lambda e: self.toggle_fullscreen())
        self.bind("<Escape>", lambda e: self.exit_fullscreen_if_any())

    def build_header(self):
        hdr = tk.Frame(self, bg=COLOR_PRIMARY, height=54)
        hdr.pack(fill='x', side='top')
        lbl = tk.Label(hdr, text=APP_TITLE, bg=COLOR_PRIMARY, fg='white', font=("Segoe UI", 12, "bold"))
        lbl.pack(side='left', padx=12)
        self.lbl_greet = tk.Label(hdr, text=f"Selamat datang, {self.user_email.split('@')[0]}", bg=COLOR_PRIMARY, fg='white', font=("Segoe UI", 10))
        self.lbl_greet.pack(side='right', padx=12)
        btn_frame = tk.Frame(hdr, bg=COLOR_PRIMARY)
        btn_frame.pack(side='right', padx=6)
        tk.Button(btn_frame, text="⛶", bg=COLOR_PRIMARY, fg='white', bd=0, command=self.toggle_fullscreen).pack(side='left', padx=6)
        tk.Button(btn_frame, text="Jalankan Siklus", bg=COLOR_ACCENT, fg='white', bd=0, command=self.run_cycle).pack(side='left', padx=6)
        tk.Button(btn_frame, text="Logout", bg=COLOR_PRIMARY, fg='white', bd=0, command=self.logout).pack(side='left', padx=6)
        tk.Button(btn_frame, text="✕", bg=COLOR_PRIMARY, fg='white', bd=0, command=self.quit_app).pack(side='left', padx=6)
        hdr.bind("<ButtonPress-1>", self.start_move)
        hdr.bind("<B1-Motion>", self.do_move)
        self._offsetx = 0
        self._offsety = 0

    def start_move(self, event):
        self._offsetx = event.x
        self._offsety = event.y

    def do_move(self, event):
        if not self._is_fullscreen:
            x = self.winfo_x() + event.x - self._offsetx
            y = self.winfo_y() + event.y - self._offsety
            self.geometry(f"+{x}+{y}")

    def toggle_fullscreen(self):
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        if not self._is_fullscreen:
            self.overrideredirect(True)
            self.geometry(f"{sw}x{sh}+0+0")
            self._is_fullscreen = True
        else:
            self.overrideredirect(False)
            w, h = 1000, 650
            x = (sw - w) // 2
            y = (sh - h) // 2
            self.geometry(f"{w}x{h}+{x}+{y}")
            self._is_fullscreen = False

    def exit_fullscreen_if_any(self):
        if self._is_fullscreen:
            self.toggle_fullscreen()

    def logout(self):
        if messagebox.askyesno("Logout", "Yakin logout?"):
            self.destroy()
            main()

    def quit_app(self):
        if messagebox.askyesno("Keluar", "Keluar aplikasi?"):
            self.destroy()

    # Notebook & tabs
    def build_notebook(self):
        self.nb = ttk.Notebook(self)
        self.nb.pack(fill='both', expand=True, padx=16, pady=(12, 18))
        # tabs
        self.tab_accounts = ttk.Frame(self.nb)
        self.tab_journal = ttk.Frame(self.nb)
        self.tab_ledger = ttk.Frame(self.nb)
        self.tab_trial = ttk.Frame(self.nb)           # Neraca Saldo (sebelum)
        self.tab_adjust = ttk.Frame(self.nb)          # Jurnal Penyesuaian
        self.tab_trial_adj = ttk.Frame(self.nb)       # Neraca Saldo setelah penyesuaian
        self.tab_reports = ttk.Frame(self.nb)         # Laporan Keuangan
        self.tab_dashboard = ttk.Frame(self.nb)
        # Order
        self.nb.add(self.tab_accounts, text="Akun")
        self.nb.add(self.tab_journal, text="Jurnal")
        self.nb.add(self.tab_ledger, text="Buku Besar")
        self.nb.add(self.tab_trial, text="Neraca Saldo")
        self.nb.add(self.tab_adjust, text="Penyesuaian")
        self.nb.add(self.tab_trial_adj, text="Neraca Setelah Penyesuaian")
        self.nb.add(self.tab_reports, text="Laporan Keuangan")
        self.nb.add(self.tab_dashboard, text="Dashboard")
        # build each
        self.build_accounts_tab()
        self.build_journal_tab()
        self.build_ledger_tab()
        self.build_trial_tab()
        self.build_adjust_tab()
        self.build_trial_adj_tab()
        self.build_reports_tab()
        self.build_dashboard_tab()

    # Accounts tab
    def build_accounts_tab(self):
        top = ttk.Frame(self.tab_accounts)
        top.pack(fill='x', pady=8)
        ttk.Label(top, text="Manajemen Akun", style='Header.TLabel').pack(side='left', padx=8)
        ttk.Button(top, text="Tambah Akun", command=self.ui_add_account).pack(side='right', padx=6)
        ttk.Button(top, text="Edit Akun", command=self.ui_edit_account).pack(side='right', padx=6)
        ttk.Button(top, text="Hapus Akun", command=self.ui_delete_account).pack(side='right', padx=6)
        cols = ("No Akun", "Nama Akun", "Tipe")
        self.tree_accounts = ttk.Treeview(self.tab_accounts, columns=cols, show='headings', height=18)
        for c in cols:
            self.tree_accounts.heading(c, text=c)
            self.tree_accounts.column(c, width=200 if c != "Nama Akun" else 420)
        self.tree_accounts.pack(fill='both', expand=True, padx=12, pady=8)

    def ui_add_account(self):
        d = AccountDialog(self)
        self.wait_window(d)
        if d.result:
            no, nama, tipe = d.result
            try:
                add_account_db(no, nama, tipe)
                messagebox.showinfo("Sukses", "Akun ditambahkan")
                self.refresh_all()
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def ui_edit_account(self):
        sel = self.tree_accounts.selection()
        if not sel:
            messagebox.showinfo("Info", "Pilih akun untuk diedit")
            return
        vals = self.tree_accounts.item(sel[0])['values']
        no = vals[0]
        d = AccountDialog(self, initial=vals, edit_mode=True)
        self.wait_window(d)
        if d.result:
            _, nama, tipe = d.result
            try:
                edit_account_db(no, nama, tipe)
                messagebox.showinfo("Sukses", "Akun diupdate")
                self.refresh_all()
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def ui_delete_account(self):
        sel = self.tree_accounts.selection()
        if not sel:
            messagebox.showinfo("Info", "Pilih akun")
            return
        vals = self.tree_accounts.item(sel[0])['values']
        no = vals[0]
        if messagebox.askyesno("Konfirmasi", f"Hapus akun {no} - {vals[1]} ?"):
            try:
                delete_account_db(no)
                messagebox.showinfo("Sukses", "Akun dihapus")
                self.refresh_all()
            except Exception as e:
                messagebox.showerror("Error", str(e))

    # Journal tab
    def build_journal_tab(self):
        top = ttk.Frame(self.tab_journal)
        top.pack(fill='x', pady=6)
        ttk.Label(top, text="Jurnal Umum", style='Header.TLabel').pack(side='left', padx=8)
        ttk.Button(top, text="Tambah Jurnal", command=self.ui_add_journal).pack(side='right', padx=6)
        ttk.Button(top, text="Hapus Jurnal", command=self.ui_delete_journal).pack(side='right', padx=6)
        ttk.Button(top, text="Export Jurnal ke Excel", command=self.export_journal_excel).pack(side='right', padx=6)
        cols = ("ID", "Tanggal", "Akun Debit", "Akun Kredit", "Debit", "Kredit", "Keterangan")
        self.tree_journal = ttk.Treeview(self.tab_journal, columns=cols, show='headings', height=18)
        for c in cols:
            self.tree_journal.heading(c, text=c)
            self.tree_journal.column(c, width=120 if c != "Keterangan" else 360)
        self.tree_journal.pack(fill='both', expand=True, padx=12, pady=8)

    def ui_add_journal(self):
        accs = list_accounts()
        if not accs:
            messagebox.showinfo("Info", "Belum ada akun. Tambahkan akun dulu di tab Akun.")
            return
        d = JournalDialog(self, accs)
        self.wait_window(d)
        if d.result:
            tanggal, debit_acc, credit_acc, debit_val, credit_val, ket = d.result
            try:
                add_journal_db(tanggal, debit_acc, credit_acc, debit_val, credit_val, ket)
                messagebox.showinfo("Sukses", "Jurnal disimpan")
                self.refresh_all()
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def ui_delete_journal(self):
        sel = self.tree_journal.selection()
        if not sel:
            messagebox.showinfo("Info", "Pilih jurnal untuk dihapus")
            return
        ids = [self.tree_journal.item(s)['values'][0] for s in sel]
        for _id in ids:
            delete_journal_db(_id)
        messagebox.showinfo("Sukses", f"{len(ids)} jurnal dihapus")
        self.refresh_all()

    # Ledger tab
    def build_ledger_tab(self):
        top = ttk.Frame(self.tab_ledger)
        top.pack(fill='x', pady=6)
        ttk.Label(top, text="Buku Besar (Pilih Akun)", style='Header.TLabel').pack(side='left', padx=8)
        self.combo_ledger_accounts = ttk.Combobox(top, values=[f"{a[0]} - {a[1]}" for a in list_accounts()], width=32)
        self.combo_ledger_accounts.pack(side='left', padx=8)
        ttk.Button(top, text="Tampilkan", command=self.refresh_ledger_for_account).pack(side='left', padx=6)
        ttk.Button(top, text="Refresh Akun", command=self.refresh_all).pack(side='right', padx=8)
        cols = ("Tanggal", "Journal ID", "Memo", "Debit", "Kredit", "Saldo Berjalan")
        self.tree_ledger_account = ttk.Treeview(self.tab_ledger, columns=cols, show='headings', height=18)
        for c in cols:
            self.tree_ledger_account.heading(c, text=c)
            self.tree_ledger_account.column(c, width=140 if c != "Memo" else 360)
        self.tree_ledger_account.pack(fill='both', expand=True, padx=12, pady=8)

    def refresh_ledger_for_account(self):
        sel = self.combo_ledger_accounts.get()
        if not sel:
            messagebox.showinfo("Info", "Pilih akun")
            return
        no = sel.split(" - ", 1)[0].strip()
        rows = []
        for r in list_journal_entries():
            _id, tanggal, ad, ak, d, k, ket = r
            if ad == no or ak == no:
                rows.append(r)
        for r in self.tree_ledger_account.get_children():
            self.tree_ledger_account.delete(r)
        bal = to_decimal('0')
        for _id, tanggal, ad, ak, d, k, ket in rows:
            if ad == no:
                bal += to_decimal(d)
                debit = f"{d:,.2f}"
                kredit = ""
            else:
                bal -= to_decimal(k)
                debit = ""
                kredit = f"{k:,.2f}"
            self.tree_ledger_account.insert("", "end", values=(tanggal, _id, ket, debit, kredit, f"{bal:,.2f}"))

    # Trial tab (before penyesuaian)
    def build_trial_tab(self):
        top = ttk.Frame(self.tab_trial)
        top.pack(fill='x', pady=6)
        ttk.Label(top, text="Neraca Saldo (Sebelum Penyesuaian)", style='Header.TLabel').pack(side='left', padx=8)
        ttk.Button(top, text="Export Neraca ke Excel", command=lambda: self.export_trial_excel(False)).pack(side='right', padx=8)
        cols = ("No Akun", "Nama Akun", "Tipe", "Debit", "Kredit")
        self.tree_trial = ttk.Treeview(self.tab_trial, columns=cols, show='headings', height=18)
        for c in cols:
            self.tree_trial.heading(c, text=c)
            self.tree_trial.column(c, width=140 if c != "Nama Akun" else 360)
        self.tree_trial.pack(fill='both', expand=True, padx=12, pady=8)

    # Penyesuaian tab
    def build_adjust_tab(self):
        top = ttk.Frame(self.tab_adjust)
        top.pack(fill='x', pady=6)
        ttk.Label(top, text="Jurnal Penyesuaian", style='Header.TLabel').pack(side='left', padx=8)
        ttk.Button(top, text="Tambah Penyesuaian", command=self.ui_add_adjust).pack(side='right', padx=6)
        ttk.Button(top, text="Hapus Penyesuaian", command=self.ui_delete_adjust).pack(side='right', padx=6)
        ttk.Button(top, text="Apply Penyesuaian (copy ke Jurnal)", command=self.ui_apply_adjustments).pack(side='right', padx=6)
        ttk.Button(top, text="Export Penyesuaian ke Excel", command=self.export_adjust_excel).pack(side='right', padx=6)
        cols = ("ID", "Tanggal", "Akun Debit", "Akun Kredit", "Debit", "Kredit", "Keterangan")
        self.tree_adjust = ttk.Treeview(self.tab_adjust, columns=cols, show='headings', height=18)
        for c in cols:
            self.tree_adjust.heading(c, text=c)
            self.tree_adjust.column(c, width=120 if c != "Keterangan" else 360)
        self.tree_adjust.pack(fill='both', expand=True, padx=12, pady=8)

    def ui_add_adjust(self):
        accs = list_accounts()
        if not accs:
            messagebox.showinfo("Info", "Belum ada akun. Tambahkan akun dulu di tab Akun.")
            return
        d = AdjustDialog(self, accs)
        self.wait_window(d)
        if d.result:
            tanggal, debit_acc, credit_acc, debit_val, credit_val, ket = d.result
            try:
                add_adjusting_db(tanggal, debit_acc, credit_acc, debit_val, credit_val, ket)
                messagebox.showinfo("Sukses", "Penyesuaian disimpan")
                self.refresh_all()
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def ui_delete_adjust(self):
        sel = self.tree_adjust.selection()
        if not sel:
            messagebox.showinfo("Info", "Pilih penyesuaian untuk dihapus")
            return
        ids = [self.tree_adjust.item(s)['values'][0] for s in sel]
        for _id in ids:
            delete_adjusting_db(_id)
        messagebox.showinfo("Sukses", f"{len(ids)} penyesuaian dihapus")
        self.refresh_all()

    def ui_apply_adjustments(self):
        cnt = apply_adjustments()
        messagebox.showinfo("Selesai", f"{cnt} penyesuaian diterapkan ke Jurnal (tanggal hari ini).")
        self.refresh_all()

    # Trial adjusted tab (after penyesuaian, virtual - includes adjusting entries)
    def build_trial_adj_tab(self):
        top = ttk.Frame(self.tab_trial_adj)
        top.pack(fill='x', pady=6)
        ttk.Label(top, text="Neraca Saldo Setelah Penyesuaian", style='Header.TLabel').pack(side='left', padx=8)
        ttk.Button(top, text="Export Neraca Penyesuaian ke Excel", command=lambda: self.export_trial_excel(True)).pack(side='right', padx=8)
        cols = ("No Akun", "Nama Akun", "Tipe", "Debit", "Kredit")
        self.tree_trial_adj = ttk.Treeview(self.tab_trial_adj, columns=cols, show='headings', height=18)
        for c in cols:
            self.tree_trial_adj.heading(c, text=c)
            self.tree_trial_adj.column(c, width=140 if c != "Nama Akun" else 360)
        self.tree_trial_adj.pack(fill='both', expand=True, padx=12, pady=8)

    # Reports tab (Laporan Keuangan)
    def build_reports_tab(self):
        top = ttk.Frame(self.tab_reports)
        top.pack(fill='x', pady=6)
        ttk.Label(top, text="Laporan Keuangan (Setelah Penyesuaian)", style='Header.TLabel').pack(side='left', padx=8)
        ttk.Button(top, text="Export Laporan ke Excel", command=self.export_reports_excel).pack(side='right', padx=6)
        ttk.Button(top, text="Export Laporan ke PDF", command=self.export_reports_pdf).pack(side='right', padx=6)
        frame = ttk.Frame(self.tab_reports)
        frame.pack(fill='both', expand=True, padx=12, pady=8)
        left = ttk.Frame(frame)
        left.pack(side='left', fill='both', expand=True, padx=6)
        mid = ttk.Frame(frame)
        mid.pack(side='left', fill='both', expand=True, padx=6)
        right = ttk.Frame(frame)
        right.pack(side='left', fill='both', expand=True, padx=6)

        # Income statement tree
        ttk.Label(left, text="Laporan Laba Rugi", style='Header.TLabel').pack(anchor='nw')
        cols = ("Tipe", "No Akun", "Nama Akun", "Jumlah")
        self.tree_income = ttk.Treeview(left, columns=cols, show='headings', height=12)
        for c in cols:
            self.tree_income.heading(c, text=c)
            self.tree_income.column(c, width=120 if c != "Nama Akun" else 220)
        self.tree_income.pack(fill='both', expand=True, pady=6)

        # Equity tree
        ttk.Label(mid, text="Perubahan Ekuitas (Rekonsiliasi)", style='Header.TLabel').pack(anchor='nw')
        cols2 = ("Keterangan", "Jumlah")
        self.tree_equity = ttk.Treeview(mid, columns=cols2, show='headings', height=8)
        for c in cols2:
            self.tree_equity.heading(c, text=c)
            self.tree_equity.column(c, width=220)
        self.tree_equity.pack(fill='both', expand=True, pady=6)

        # Balance sheet tree
        ttk.Label(right, text="Neraca (Balance Sheet)", style='Header.TLabel').pack(anchor='nw')
        cols3 = ("Sisi", "No Akun", "Nama Akun", "Jumlah")
        self.tree_balance = ttk.Treeview(right, columns=cols3, show='headings', height=20)
        for c in cols3:
            self.tree_balance.heading(c, text=c)
            self.tree_balance.column(c, width=100 if c != "Nama Akun" else 200)
        self.tree_balance.pack(fill='both', expand=True, pady=6)

    # Dashboard
    def build_dashboard_tab(self):
        top = ttk.Frame(self.tab_dashboard)
        top.pack(fill='x', pady=6)
        ttk.Label(top, text="Dashboard", style='Header.TLabel').pack(side='left', padx=8)
        ttk.Button(top, text="Refresh", command=self.refresh_all).pack(side='right', padx=8)
        cards_frame = ttk.Frame(self.tab_dashboard)
        cards_frame.pack(fill='both', expand=True, padx=16, pady=12)
        self.card1 = tk.Frame(cards_frame, bg=CARD_BG, bd=1, relief='flat')
        self.card1.pack(side='left', fill='both', expand=True, padx=8, pady=8)
        tk.Label(self.card1, text="Aset vs Kewajiban", bg=CARD_BG, fg=COLOR_PRIMARY, font=("Segoe UI", 12, "bold")).pack(anchor='nw', padx=12, pady=8)
        self.canvas1 = None
        self.card2 = tk.Frame(cards_frame, bg=CARD_BG, bd=1, relief='flat')
        self.card2.pack(side='left', fill='both', expand=True, padx=8, pady=8)
        tk.Label(self.card2, text="Likuiditas (Current Ratio)", bg=CARD_BG, fg=COLOR_PRIMARY, font=("Segoe UI", 12, "bold")).pack(anchor='nw', padx=12, pady=8)
        self.canvas2 = None
        self.card3 = tk.Frame(cards_frame, bg=CARD_BG, bd=1, relief='flat')
        self.card3.pack(side='left', fill='both', expand=True, padx=8, pady=8)
        tk.Label(self.card3, text="Solvabilitas (Debt-to-Equity)", bg=CARD_BG, fg=COLOR_PRIMARY, font=("Segoe UI", 12, "bold")).pack(anchor='nw', padx=12, pady=8)
        self.canvas3 = None
        if plt and FigureCanvasTkAgg:
            self.fig1 = plt.Figure(figsize=(4, 2))
            self.ax1 = self.fig1.add_subplot(111)
            self.canvas1 = FigureCanvasTkAgg(self.fig1, master=self.card1)
            self.canvas1.get_tk_widget().pack(fill='both', expand=True, padx=8, pady=6)
            self.fig2 = plt.Figure(figsize=(4, 2))
            self.ax2 = self.fig2.add_subplot(111)
            self.canvas2 = FigureCanvasTkAgg(self.fig2, master=self.card2)
            self.canvas2.get_tk_widget().pack(fill='both', expand=True, padx=8, pady=6)
            self.fig3 = plt.Figure(figsize=(4, 2))
            self.ax3 = self.fig3.add_subplot(111)
            self.canvas3 = FigureCanvasTkAgg(self.fig3, master=self.card3)
            self.canvas3.get_tk_widget().pack(fill='both', expand=True, padx=8, pady=6)
        else:
            tk.Label(self.card1, text="matplotlib tidak tersedia", bg=CARD_BG).pack()
            tk.Label(self.card2, text="matplotlib tidak tersedia", bg=CARD_BG).pack()
            tk.Label(self.card3, text="matplotlib tidak tersedia", bg=CARD_BG).pack()

    # Refresh all
    def refresh_all(self):
        # accounts
        for r in self.tree_accounts.get_children():
            self.tree_accounts.delete(r)
        for no, nm, tipe in list_accounts():
            self.tree_accounts.insert("", "end", values=(no, nm, tipe))
        # journal
        for r in self.tree_journal.get_children():
            self.tree_journal.delete(r)
        for id_, tgl, ad, ak, d, k, ket in list_journal_entries():
            self.tree_journal.insert("", "end", values=(id_, tgl, ad, ak, f"{d:,.2f}", f"{k:,.2f}", ket))
        # adjusting
        if hasattr(self, 'tree_adjust'):
            for r in self.tree_adjust.get_children():
                self.tree_adjust.delete(r)
            for id_, tgl, ad, ak, d, k, ket in list_adjusting_entries():
                self.tree_adjust.insert("", "end", values=(id_, tgl, ad, ak, f"{d:,.2f}", f"{k:,.2f}", ket))
        # trial (before)
        for r in self.tree_trial.get_children():
            self.tree_trial.delete(r)
        for no, nm, tipe, d, c in compute_trial_rows():
            self.tree_trial.insert("", "end", values=(no, nm, tipe, f"{d:,.2f}", f"{c:,.2f}"))
        # trial adjusted (virtual include adjusting entries)
        for r in self.tree_trial_adj.get_children():
            self.tree_trial_adj.delete(r)
        for no, nm, tipe, d, c in compute_trial_rows(include_adjustments=True):
            self.tree_trial_adj.insert("", "end", values=(no, nm, tipe, f"{d:,.2f}", f"{c:,.2f}"))
        # ledger combo
        if hasattr(self, 'combo_ledger_accounts'):
            self.combo_ledger_accounts['values'] = [f"{a[0]} - {a[1]}" for a in list_accounts()]
        # refresh reports
        self.refresh_reports()
        # draw dashboard
        self.draw_dashboard()

    def refresh_reports(self):
        inc, bal, eqrec = compute_financial_statements(include_adjustments=True)
        # income tree
        for r in self.tree_income.get_children():
            self.tree_income.delete(r)
        for no, nm, amt in inc['revenues']:
            self.tree_income.insert("", "end", values=("Pendapatan", no, nm, f"{amt:,.2f}"))
        self.tree_income.insert("", "end", values=("", "", "Total Pendapatan", f"{inc['total_revenue']:,.2f}"))
        for no, nm, amt in inc['expenses']:
            self.tree_income.insert("", "end", values=("Beban", no, nm, f"{amt:,.2f}"))
        self.tree_income.insert("", "end", values=("", "", "Total Beban", f"{inc['total_expense']:,.2f}"))
        self.tree_income.insert("", "end", values=("", "", "Laba (Rugi) Bersih", f"{inc['net_income']:,.2f}"))

        # equity
        for r in self.tree_equity.get_children():
            self.tree_equity.delete(r)
        for label, amt in eqrec:
            self.tree_equity.insert("", "end", values=(label, f"{amt:,.2f}"))

        # balance
        for r in self.tree_balance.get_children():
            self.tree_balance.delete(r)
        for no, nm, amt in bal['assets']:
            self.tree_balance.insert("", "end", values=("Aset", no, nm, f"{amt:,.2f}"))
        self.tree_balance.insert("", "end", values=("", "", "Total Aset", f"{bal['total_assets']:,.2f}"))
        for no, nm, amt in bal['liabilities']:
            self.tree_balance.insert("", "end", values=("Kewajiban", no, nm, f"{amt:,.2f}"))
        for no, nm, amt in bal['equity']:
            self.tree_balance.insert("", "end", values=("Ekuitas", no, nm, f"{amt:,.2f}"))
        self.tree_balance.insert("", "end", values=("", "", "Total Kewajiban + Ekuitas", f"{(bal['total_liabilities'] + bal['total_equity']):,.2f}"))

    def draw_dashboard(self):
        stats = prepare_balance_and_ratios(include_adjustments=False)
        aset = float(stats['assets'])
        liab = float(stats['liabilities'])
        if self.canvas1:
            self.ax1.clear()
            # DO NOT set colors explicitly outside user's request? we keep the theme colors set earlier
            self.ax1.bar(['Aset', 'Kewajiban'], [aset, liab])
            self.canvas1.draw()
        if self.canvas2:
            self.ax2.clear()
            cr = stats['current_ratio']
            if cr is None:
                self.ax2.text(0.5, 0.5, "N/A", ha='center', va='center', fontsize=14)
            else:
                val = float(cr)
                self.ax2.barh(['CR'], [val])
                self.ax2.set_xlim(0, max(1.5, val * 1.2))
            self.canvas2.draw()
        if self.canvas3:
            self.ax3.clear()
            dte = stats['debt_to_equity']
            if dte is None:
                self.ax3.text(0.5, 0.5, "N/A", ha='center', va='center', fontsize=14)
            else:
                val = float(dte)
                self.ax3.barh(['D/E'], [val])
                self.ax3.set_xlim(0, max(1.5, val * 1.2))
            self.canvas3.draw()

    # Export wrappers
    def export_trial_excel(self, include_adjustments=False):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            export_trial_to_excel(path, include_adjustments)
            messagebox.showinfo("Sukses", "Neraca disimpan.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def export_journal_excel(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            export_journal_to_excel(path)
            messagebox.showinfo("Sukses", "Jurnal disimpan.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def export_adjust_excel(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            export_adjusting_to_excel(path)
            messagebox.showinfo("Sukses", "Penyesuaian disimpan.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def export_reports_pdf(self):
        path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
        if not path:
            return
        try:
            export_reports_to_pdf(path, include_adjustments=True)
            messagebox.showinfo("Sukses", "Laporan PDF disimpan.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def export_reports_excel(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            export_reports_to_excel(path, include_adjustments=True)
            messagebox.showinfo("Sukses", "Laporan Excel disimpan.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # --- AUTOMATION: run cycle ---
    def run_cycle(self):
        """
        Automatisasi urutan siklus:
        1) Pastikan ada Jurnal (user)
        2) Terapkan semua Penyesuaian -> salin ke jurnal (apply_adjustments)
        3) Refresh Buku Besar, Neraca Saldo & Neraca Setelah Penyesuaian & Laporan Keuangan & Dashboard
        """
        if messagebox.askyesno("Jalankan Siklus", "Jalankan siklus: terapkan semua penyesuaian ke jurnal dan refresh laporan?"):
            applied = apply_adjustments()
            self.refresh_all()
            messagebox.showinfo("Selesai", f"Siklus selesai. {applied} penyesuaian diterapkan ke jurnal (tanggal hari ini).")

# ---------------- Program start ----------------
def main():
    init_db_if_needed()
    if not SMTP_EMAIL or not SMTP_PASSWORD:
        print("NOTE: SMTP kosong - OTP tidak terkirim email; OTP ditampilkan di console saat register.")
    missing = []
    if pd is None:
        missing.append("pandas")
    if matplotlib is None or plt is None:
        missing.append("matplotlib")
    if Workbook is None:
        missing.append("openpyxl")
    if SimpleDocTemplate is None:
        missing.append("reportlab")
    if missing:
        print("Optional modules missing:", ", ".join(missing))

    root = tk.Tk()
    root.withdraw()

    def on_success(email):
        root.destroy()
        app = MainApp(email)
        app.mainloop()

    login = LoginWindow(root, on_success)
    login.grab_set()
    root.mainloop()

if __name__ == "__main__":
    main()
